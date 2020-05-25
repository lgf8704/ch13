import docx
import openpyxl


def get_names(excel_file_name, sheet='Sheet1'):
    wb = openpyxl.load_workbook(excel_file_name)
    ws = wb['Sheet1']
    names = []
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        names.append(cell.value)

    return names


def write_doc(doc_name):
    doc = docx.Document()
    for name in names:
        text = """It would be a pleasure to have the company of 
                            Robocop 
                at 11010 Memory Love at the evening of 
                        April 1st
                    at 7 o'clock
        """
    doc.save(doc_name)




if __name__ == "__main__":
    excel_file_name = "17学生基本信息表.xls"
    names = get_names(excel_file_name)

    doc_name = "invite_doc.docx"
    write_doc(doc_name)